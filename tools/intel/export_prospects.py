#!/usr/bin/env python3
"""
tools/intel/export_prospects.py — 客户线索导出工具

将 customer_prospect 表数据导出为 Excel（.xlsx）或 CSV。

用法:
  # 导出指定 brand_id 的全部线索
  python3 tools/intel/export_prospects.py --brand-id EXPO-0001

  # 导出指定 report_id 的线索
  python3 tools/intel/export_prospects.py --report-id 3

  # 导出所有线索（CSV）
  python3 tools/intel/export_prospects.py --all --format csv

  # 指定输出路径
  python3 tools/intel/export_prospects.py --brand-id EXPO-0001 --out /tmp/leads.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
DB_PATH = _REPO_ROOT / "data" / "mwlab.db"

_COLUMNS = [
    "id", "intel_report_id", "brand_id", "source_type",
    "qcc_key_no", "company_name", "credit_code", "oper_name",
    "start_date", "company_status", "reg_no", "address",
    "prospect_score", "contact_status", "notes",
    "created_at", "updated_at",
]

_COL_LABELS = {
    "id": "ID",
    "intel_report_id": "报告ID",
    "brand_id": "展会品牌",
    "source_type": "来源",
    "qcc_key_no": "企查查KeyNo",
    "company_name": "公司名称",
    "credit_code": "统一社会信用代码",
    "oper_name": "法定代表人",
    "start_date": "成立日期",
    "company_status": "企业状态",
    "reg_no": "注册号",
    "address": "注册地址",
    "prospect_score": "潜力评分(1-5)",
    "contact_status": "接触状态",
    "notes": "备注",
    "created_at": "创建时间",
    "updated_at": "更新时间",
}


def _query_prospects(
    brand_id: str | None = None,
    report_id: int | None = None,
    all_records: bool = False,
    db_path: str | Path | None = None,
) -> list[dict]:
    target_db = db_path or DB_PATH
    conn = sqlite3.connect(str(target_db))
    conn.row_factory = sqlite3.Row

    if brand_id:
        rows = conn.execute(
            f"SELECT {', '.join(_COLUMNS)} FROM customer_prospect WHERE brand_id = ? ORDER BY id",
            (brand_id,),
        ).fetchall()
    elif report_id is not None:
        rows = conn.execute(
            f"SELECT {', '.join(_COLUMNS)} FROM customer_prospect WHERE intel_report_id = ? ORDER BY id",
            (report_id,),
        ).fetchall()
    elif all_records:
        rows = conn.execute(
            f"SELECT {', '.join(_COLUMNS)} FROM customer_prospect ORDER BY id",
        ).fetchall()
    else:
        conn.close()
        raise ValueError("必须提供 --brand-id、--report-id 或 --all 之一")

    result = [dict(r) for r in rows]
    conn.close()
    return result


def _default_output_path(fmt: str, brand_id: str | None, report_id: int | None) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = f"brand_{brand_id}" if brand_id else (f"report_{report_id}" if report_id else "all")
    filename = f"prospects_{slug}_{ts}.{fmt}"
    out_dir = _REPO_ROOT / "reports" / "customer"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir / filename


def export_xlsx(rows: list[dict], out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("错误: 需要安装 openpyxl（pip install openpyxl）", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户线索"

    # 标题行
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="DDEEFF")
    headers = [_COL_LABELS.get(c, c) for c in _COLUMNS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # 数据行
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(_COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col))

    # 列宽自适应（粗略估算）
    col_widths = {
        "company_name": 30, "address": 40, "notes": 35,
        "credit_code": 22, "qcc_key_no": 20,
    }
    for col_idx, col in enumerate(_COLUMNS, 1):
        width = col_widths.get(col, 15)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(str(out_path))


def export_csv(rows: list[dict], out_path: Path) -> None:
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=_COLUMNS)
        writer.writerow({c: _COL_LABELS.get(c, c) for c in _COLUMNS})
        writer.writerows(rows)


def export_prospects(
    brand_id: str | None = None,
    report_id: int | None = None,
    all_records: bool = False,
    fmt: str = "xlsx",
    out_path: Path | None = None,
    db_path: str | Path | None = None,
) -> Path:
    """查询并导出线索，返回输出文件路径。"""
    rows = _query_prospects(brand_id=brand_id, report_id=report_id, all_records=all_records, db_path=db_path)
    if not rows:
        raise ValueError("未找到符合条件的客户线索记录")

    if out_path is None:
        out_path = _default_output_path(fmt, brand_id, report_id)

    if fmt == "xlsx":
        export_xlsx(rows, out_path)
    else:
        export_csv(rows, out_path)

    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="客户线索导出工具（Excel/CSV）")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--brand-id", help="按展会品牌 ID 过滤")
    group.add_argument("--report-id", type=int, help="按 intel_report_id 过滤")
    group.add_argument("--all", action="store_true", dest="all_records", help="导出全部记录")
    parser.add_argument("--format", default=None, choices=["xlsx", "csv"], help="输出格式（默认从 --out 扩展名推断）")
    parser.add_argument("--out", help="输出文件路径（可选，默认 reports/customer/）")
    parser.add_argument("--db", default=str(DB_PATH), help="数据库路径（默认 mwlab.db）")
    args = parser.parse_args()

    fmt = args.format
    out_path = Path(args.out) if args.out else None

    # 从 --out 扩展名推断格式，与 --format 冲突时报错
    if out_path and out_path.suffix in ('.xlsx', '.csv'):
        ext_fmt = 'xlsx' if out_path.suffix == '.xlsx' else 'csv'
        if fmt and fmt != ext_fmt:
            print(f"错误: --out 扩展名 ({out_path.suffix}) 与 --format ({fmt}) 冲突", file=sys.stderr)
            sys.exit(1)
        fmt = ext_fmt

    fmt = fmt or 'xlsx'  # 默认 xlsx

    try:
        result_path = export_prospects(
            brand_id=args.brand_id,
            report_id=args.report_id,
            all_records=args.all_records,
            fmt=fmt,
            out_path=out_path,
            db_path=args.db,
        )
        print(f"导出成功 → {result_path}")
    except ValueError as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
