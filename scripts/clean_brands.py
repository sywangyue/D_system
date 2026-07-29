#!/usr/bin/env python3
"""
clean_brands.py — Phase 05 品牌表清洗

Usage:
    python scripts/clean_brands.py name-en            # CLEAN-NAME-EN
    python scripts/clean_brands.py industry            # CLEAN-INDUSTRY
    python scripts/clean_brands.py mds                 # CLEAN-MDS (stub)
    python scripts/clean_brands.py jufair-l2           # CLEAN-JUFAIR-L2 (stub)
    python scripts/clean_brands.py --dry-run name-en   # Preview only, no writes
    python scripts/clean_brands.py --db /path/to/mwlab.db name-en
"""
from __future__ import annotations

import argparse
import difflib
import logging
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from uuid import uuid4

# Ensure project root is on sys.path for imports like scripts.data.*
BASE_DIR = Path(__file__).parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

DB_PATH = BASE_DIR / "data" / "mwlab.db"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


# ─── Utilities ────────────────────────────────────────────────────────────────

def backup_table(conn: sqlite3.Connection) -> None:
    """在 UPDATE 前创建 exhibition_brand 表的快照。

    快照表名: exhibition_brand_backup_YYYYMMDD
    如果当天快照已存在则跳过（幂等）。
    """
    today = datetime.now().strftime("%Y%m%d")
    backup_name = f"exhibition_brand_backup_{today}"
    exists = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (backup_name,),
    ).fetchone()
    if exists:
        log.info("备份表 %s 已存在，跳过", backup_name)
        return
    conn.execute(f"CREATE TABLE {backup_name} AS SELECT * FROM exhibition_brand")
    row_count = conn.execute(f"SELECT COUNT(*) FROM {backup_name}").fetchone()[0]
    log.info("已创建备份表 %s（%d 行）", backup_name, row_count)


# ─── name-en 子命令 ───────────────────────────────────────────────────────────

def cmd_name_en(args: argparse.Namespace) -> None:
    """CLEAN-NAME-EN: 标准化英文名称。

    Step 1: 清除 name_en 中包含中文字符的行（置为空）
    Step 2: 从 name_cn 提取嵌入英文（extract_embedded_en）
    Step 3: 为剩余空行生成标准英文名（generate_name_en）
    """
    from scripts.data.name_en_patterns import extract_embedded_en, generate_name_en

    conn = sqlite3.connect(str(args.db))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row

    if not args.dry_run:
        backup_table(conn)

    total = conn.execute("SELECT COUNT(*) FROM exhibition_brand").fetchone()[0]
    log.info("总行数: %d", total)

    # Step 1: 清除含中文的 name_en
    log.info("Step 1: 清除含有中文字符的 name_en...")
    if args.dry_run:
        count_dirty = conn.execute(
            "SELECT COUNT(*) FROM exhibition_brand WHERE name_en GLOB '*[一-龥]*'"
        ).fetchone()[0]
        log.info("[DRY-RUN] 将清除 %d 行含中文的 name_en", count_dirty)
        cleared = 0
    else:
        conn.execute(
            "UPDATE exhibition_brand SET name_en = '' WHERE name_en GLOB '*[一-龥]*'"
        )
        cleared = conn.execute("SELECT changes()").fetchone()[0]
    log.info("Step 1 完成: 处理 %d 行", cleared)

    # Step 2: 从 name_cn 提取嵌入英文
    log.info("Step 2: 从 name_cn 提取嵌入英文...")
    rows_empty_en = conn.execute(
        "SELECT brand_id, name_cn FROM exhibition_brand WHERE name_en = '' OR name_en IS NULL"
    ).fetchall()
    extracted = 0
    for row in rows_empty_en:
        en = extract_embedded_en(row["name_cn"] or "")
        if en:
            if args.dry_run:
                log.info("  [DRY-RUN] %s: name_cn=%r -> name_en=%r",
                         row["brand_id"], row["name_cn"], en)
            else:
                conn.execute(
                    "UPDATE exhibition_brand SET name_en = ? WHERE brand_id = ?",
                    (en, row["brand_id"]),
                )
            extracted += 1
    log.info("Step 2 完成: 提取 %d 个嵌入英文名", extracted)

    # Step 3: 为剩余空行生成英文名
    log.info("Step 3: 为剩余空行生成英文名...")
    rows_remaining = conn.execute(
        "SELECT brand_id, name_cn FROM exhibition_brand WHERE name_en = '' OR name_en IS NULL"
    ).fetchall()
    generated = 0
    for row in rows_remaining:
        en = generate_name_en(row["name_cn"] or "")
        if en:
            if args.dry_run:
                log.info("  [DRY-RUN] %s: name_cn=%r -> name_en=%r",
                         row["brand_id"], row["name_cn"], en)
            else:
                conn.execute(
                    "UPDATE exhibition_brand SET name_en = ? WHERE brand_id = ?",
                    (en, row["brand_id"]),
                )
            generated += 1
    log.info("Step 3 完成: 生成 %d 个英文名", generated)

    # 统计
    remaining = conn.execute(
        "SELECT COUNT(*) FROM exhibition_brand WHERE name_en = '' OR name_en IS NULL"
    ).fetchone()[0]
    skipped = total - cleared - extracted - generated - remaining
    log.info(
        "统计: 总行数=%d 清除含中文=%d 提取=%d 生成=%d 剩余=%d 跳过=%d",
        total, cleared, extracted, generated, remaining, skipped,
    )

    if args.dry_run:
        log.info("DRY-RUN 模式: 未写库，全部回滚")
    else:
        conn.commit()
        log.info("已提交变更到数据库")
    conn.close()


# ─── industry 子命令 ─────────────────────────────────────────────────────────

def cmd_industry(args: argparse.Namespace) -> None:
    """CLEAN-INDUSTRY: 将 industry_l1 归并为 6 个 MD 类别。

    Step 1: 查询所有 industry_l1 != '' 的行
    Step 2: 对每一行调用 classify_industry_l1 映射到目标类别
    Step 3: UPDATE 表（target 非空且与原始值不同时）
    Step 4: 打印映射统计和未匹配列表
    """
    from scripts.data.md_category_rules import classify_industry_l1, list_nonempty_categories

    conn = sqlite3.connect(str(args.db))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row

    if not args.dry_run:
        backup_table(conn)

    categories = list_nonempty_categories()
    log.info("MD 类别: %s", categories)

    rows = conn.execute(
        "SELECT brand_id, name_cn, industry_l1 FROM exhibition_brand "
        "WHERE industry_l1 != ''"
    ).fetchall()
    total_processed = len(rows)
    log.info("待处理行数: %d", total_processed)

    mapped_count = 0
    unchanged_count = 0
    unmapped: set[str] = set()

    for row in rows:
        original = (row["industry_l1"] or "").strip()
        target = classify_industry_l1(original)

        if not target:
            unmapped.add(original)
            continue

        if target == original:
            unchanged_count += 1
            continue

        mapped_count += 1
        if args.dry_run:
            log.info("  [DRY-RUN] %s (%s): %r -> %r",
                     row["brand_id"], row["name_cn"], original, target)
        else:
            conn.execute(
                "UPDATE exhibition_brand SET industry_l1 = ? WHERE brand_id = ?",
                (target, row["brand_id"]),
            )

    # 统计
    log.info(
        "映射统计: 总处理=%d 已映射=%d 未变更=%d 未匹配=%d",
        total_processed, mapped_count, unchanged_count, len(unmapped),
    )

    if unmapped:
        log.warning("未匹配的 industry_l1 值（共 %d 个）:", len(unmapped))
        for val in sorted(unmapped):
            log.warning("  - %s", val)

    if args.dry_run:
        log.info("DRY-RUN 模式: 未写库，全部回滚")
    else:
        conn.commit()
        log.info("已提交变更到数据库")
    conn.close()


# ─── mds 辅助函数 ────────────────────────────────────────────────────────────

def _extract_english_trailing(text: str) -> str:
    """从中英文混合字符串末尾提取英文部分。

    取最后一个中文字符之后的所有内容。
    如果字符串不含中文字符，返回原字符串（视为纯英文名）。
    """
    if not text:
        return ""
    text = text.replace("\n", " ").replace("\r", " ").strip()
    cjk_matches = list(re.finditer(r'[一-鿿]', text))
    if cjk_matches:
        trailing = text[cjk_matches[-1].end():].strip()
        return trailing
    # No CJK characters — entire string is likely English
    return text


def parse_md_excel(filepath: str | Path) -> list[dict]:
    """解析杜塞境外展 Excel 文件。

    处理合并单元格：跟踪当前 category（列 B）和 parent exhibition（列 C）。
    列布局（verified from actual file）:
      B: 类别, C: 杜塞全球展会（母展）, E: 卫星展-CN,
      F: 卫星展-EN, G: 地点, H: 展会日期（下届）

    Returns:
        list of {category, parent_cn, parent_en, sat_cn, sat_en, location, next_date}
    """
    import openpyxl  # noqa: PLC0415 — conditional import, not always needed

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    records = []
    current_cat = ""
    current_parent = ""

    for row in range(4, ws.max_row + 1):
        b_val = ws.cell(row, 2).value  # 类别
        c_val = ws.cell(row, 3).value  # 母展
        e_val = ws.cell(row, 5).value  # 卫星展-CN
        f_val = ws.cell(row, 6).value  # 卫星展-EN
        g_val = ws.cell(row, 7).value  # 地点
        h_val = ws.cell(row, 8).value  # 下届日期

        # Update current category if non-empty
        if b_val and str(b_val).strip():
            current_cat = str(b_val).strip()

        # Update current parent if non-empty
        if c_val and str(c_val).strip():
            current_parent = str(c_val).strip()

        # Skip rows where ALL data columns are empty
        if not any([e_val, f_val, g_val, h_val]):
            continue

        # / or - means no satellite show
        sat_cn = ""
        if e_val:
            e_str = str(e_val).strip()
            if e_str not in ("", "/", "-"):
                sat_cn = e_str

        sat_en = ""
        if f_val:
            f_str = str(f_val).strip()
            if f_str not in ("", "/", "-"):
                sat_en = f_str

        records.append({
            "category": current_cat,
            "parent_cn": current_parent,
            "parent_en": _extract_english_trailing(current_parent),
            "sat_cn": sat_cn,
            "sat_en": sat_en,
            "location": str(g_val).strip() if g_val else "",
            "next_date": str(h_val).strip() if h_val else "",
        })

    return records


def match_brand_multistrategy(
    conn: sqlite3.Connection,
    name_cn: str = "",
    name_en: str = "",
    threshold: float = 0.80,
) -> str | None:
    """多策略品牌匹配。

    按优先级尝试以下匹配策略：
      Strategy 1: name_en 精确匹配
      Strategy 2: name_cn 精确匹配
      Strategy 3: name_cn LIKE 子串（前 10/8/6 字）
      Strategy 4: 主办方含杜塞尔（前置条件：搜索词含杜塞尔关键词）
      Strategy 5: difflib.SequenceMatcher 模糊匹配（阈值 threshold，最小长度 >= 6）

    Returns:
        brand_id 或 None
    """
    search_cn = name_cn.strip() if name_cn else ""
    search_en = name_en.strip() if name_en else ""

    # Strategy 1: Exact name_en match
    if search_en:
        row = conn.execute(
            "SELECT brand_id FROM exhibition_brand WHERE name_en = ?",
            (search_en,),
        ).fetchone()
        if row:
            return row[0]

    # Strategy 2: Exact name_cn match
    if search_cn:
        row = conn.execute(
            "SELECT brand_id FROM exhibition_brand WHERE name_cn = ?",
            (search_cn,),
        ).fetchone()
        if row:
            return row[0]

    # Strategy 3: name_cn LIKE substring
    if search_cn and len(search_cn) >= 4:
        for prefix_len in (10, 8, 6):
            part = search_cn[:prefix_len]
            if len(part) >= 4:
                row = conn.execute(
                    "SELECT brand_id FROM exhibition_brand "
                    "WHERE name_cn LIKE ? LIMIT 1",
                    (f"%{part}%",),
                ).fetchone()
                if row:
                    return row[0]

    # Strategy 4: Organizer contains 杜塞尔
    combined = (search_cn + " " + search_en).lower()
    if "杜塞尔" in combined or "dusseldorf" in combined:
        row = conn.execute(
            "SELECT brand_id FROM exhibition_brand "
            "WHERE organizer LIKE '%杜塞尔%' LIMIT 1"
        ).fetchone()
        if row:
            return row[0]

    # Strategy 5: difflib fuzzy match
    all_rows = conn.execute(
        "SELECT brand_id, name_cn FROM exhibition_brand"
    ).fetchall()
    search_text = search_cn or search_en
    if search_text and len(search_text) >= 6:
        best_ratio, best_id = 0.0, None
        for bid, name in all_rows:
            if not name:
                continue
            ratio = difflib.SequenceMatcher(None, search_text, name).ratio()
            if ratio > best_ratio:
                best_ratio, best_id = ratio, bid
        if best_ratio >= threshold:
            return best_id

    return None


# ─── mds 子命令 ──────────────────────────────────────────────────────────────

def cmd_mds(args: argparse.Namespace) -> None:
    """CLEAN-MDS: 从 Excel 标记 MD 自有品牌。

    Step 1: 解析 Excel → 品牌记录 list
    Step 2: 对每一条记录，多策略匹配 exhibition_brand
    Step 3: 匹配成功 → UPDATE mds_related = category
    Step 4: 未匹配的母展 → INSERT 新品牌到 exhibition_brand
    """
    conn = sqlite3.connect(str(args.db))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row

    if not args.dry_run:
        backup_table(conn)

    # Parse Excel
    excel_path = BASE_DIR / "杜塞境外展时间表_for update_2026.xlsx"
    if not excel_path.exists():
        log.error("Excel 文件不存在: %s", excel_path)
        conn.close()
        return

    records = parse_md_excel(str(excel_path))
    log.info("Excel 共 %d 条记录", len(records))

    matched = 0
    new_brands = 0
    unmatched_parents: list[dict] = []
    inserted_parents: set[str] = set()

    for rec in records:
        brand_id = None

        # Try matching satellite show first
        if rec["sat_cn"] or rec["sat_en"]:
            brand_id = match_brand_multistrategy(
                conn, name_cn=rec["sat_cn"], name_en=rec["sat_en"]
            )

        # Then try parent exhibition
        if brand_id is None:
            brand_id = match_brand_multistrategy(
                conn, name_cn=rec["parent_cn"], name_en=rec["parent_en"]
            )

        if brand_id:
            matched += 1
            if args.dry_run:
                log.info("  [DRY-RUN] %s: mds_related='%s'",
                         brand_id, rec["category"])
            else:
                conn.execute(
                    "UPDATE exhibition_brand SET mds_related = ? WHERE brand_id = ?",
                    (rec["category"], brand_id),
                )
        else:
            # Not matched — track parent for INSERT
            parent_key = rec["parent_cn"].strip()
            if parent_key and parent_key not in inserted_parents:
                inserted_parents.add(parent_key)
                unmatched_parents.append(rec)

    # Insert new brands for unmatched parent exhibitions
    for rec in unmatched_parents:
        new_id = f"EXPO-{uuid4().hex[:8].upper()}"
        parent_en = rec["parent_en"] or ""
        if not parent_en:
            from scripts.data.name_en_patterns import generate_name_en  # noqa: PLC0415

            parent_en = generate_name_en(rec["parent_cn"])

        if args.dry_run:
            log.info("  [DRY-RUN] NEW BRAND: %s (%s) -> %s",
                     new_id, rec["parent_cn"], parent_en)
        else:
            conn.execute(
                "INSERT INTO exhibition_brand "
                "(brand_id, name_cn, name_en, industry_l1, mds_related) "
                "VALUES (?, ?, ?, ?, ?)",
                (new_id, rec["parent_cn"].strip(), parent_en,
                 rec["category"], rec["category"]),
            )
            log.warning("NEW BRAND (needs review): %s -> %s (%s)",
                        new_id, rec["parent_cn"], rec["category"])
        new_brands += 1

    unmatched_count = len(records) - matched - min(len(unmatched_parents), new_brands)
    log.info(
        "统计: 总行数=%d 已匹配=%d 新品牌=%d 未匹配=%d",
        len(records), matched, new_brands, unmatched_count,
    )

    if args.dry_run:
        log.info("DRY-RUN 模式: 未写库，全部回滚")
    else:
        conn.commit()
        log.info("已提交变更到数据库")
    conn.close()


# ─── jufair-l2 子命令 ─────────────────────────────────────────────────────────

def cmd_jufair_l2(args: argparse.Namespace) -> None:
    """CLEAN-JUFAIR-L2: 爬取 jufair 分类并匹配 L2。

    --export <path>:  在 Mac Mini（大陆 IP）执行爬取，输出 JSON 分类文件
    --import <path>:  在本机导入 JSON，模糊匹配到 exhibition_brand 的 industry_l1 + industry_l2
    """
    import csv as _csv  # noqa: PLC0415
    from scripts.data.jufair_l2_crawler import (  # noqa: PLC0415
        crawl_jufair_categories,
        export_categories,
        load_categories,
    )

    # --export mode: crawl and save JSON
    if args.export and not args.import_path:
        log.info("开始爬取 jufair.com 分类...")
        data = crawl_jufair_categories()
        export_categories(data, args.export)
        log.info("爬取完成，已导出到 %s", args.export)
        print()
        print("=== Next Steps ===")
        print(f"1. 将 {args.export} 复制回本机（开发机）")
        print(f"2. 在本机运行: python scripts/clean_brands.py jufair-l2 "
              f"--import {args.export} --threshold {args.threshold}")
        return

    # --import mode: load JSON and fuzzy match to exhibition_brand
    if args.import_path:
        import_path = Path(args.import_path)
        if not import_path.exists():
            log.error("文件不存在: %s", import_path)
            return

        data = load_categories(str(import_path))
        conn = sqlite3.connect(str(args.db))
        conn.execute("PRAGMA foreign_keys = ON")
        conn.row_factory = sqlite3.Row

        if not args.dry_run:
            backup_table(conn)

        # Build subcategory list and parent name map
        subcategories = data.get("subcategories", [])
        parent_map = {
            p["parent_id"]: p["name"]
            for p in data.get("parent_categories", [])
        }
        log.info("加载 %d 个子分类，开始模糊匹配...", len(subcategories))

        if not subcategories:
            log.warning("JSON 中无子分类数据，跳过匹配")
            conn.close()
            return

        # Get all brands with name_cn
        brands = conn.execute(
            "SELECT brand_id, name_cn FROM exhibition_brand"
        ).fetchall()

        matched = 0
        needs_review: list[dict[str, str]] = []

        for brand in brands:
            name_cn = brand["name_cn"] or ""
            if not name_cn or len(name_cn) < 4:
                continue

            # Find best matching jufair subcategory
            best_match = None
            best_score = 0.0
            for sub in subcategories:
                score = difflib.SequenceMatcher(
                    None, name_cn, sub["name"]
                ).ratio()
                if score > best_score:
                    best_score = score
                    best_match = sub

            if best_match is None:
                continue

            parent_name = parent_map.get(best_match["parent_id"], "")

            if best_score >= args.threshold:
                # Good match — proceed with UPDATE
                if args.dry_run:
                    log.info(
                        "  [DRY-RUN] %s (%s): "
                        "industry_l1=%s industry_l2=%s (score=%.2f)",
                        brand["brand_id"], name_cn,
                        parent_name, best_match["name"], best_score,
                    )
                else:
                    conn.execute(
                        "UPDATE exhibition_brand SET industry_l1 = ?, "
                        "industry_l2 = ? WHERE brand_id = ?",
                        (parent_name, best_match["name"], brand["brand_id"]),
                    )
                matched += 1
            elif best_score >= 0.50:
                # Below threshold but worth review
                needs_review.append({
                    "brand_id": brand["brand_id"],
                    "name_cn": name_cn,
                    "suggested_l1": parent_name,
                    "suggested_l2": best_match["name"],
                    "score": f"{best_score:.2f}",
                })

        log.info(
            "匹配统计: 总品牌数=%d 已匹配=%d 需人工复核=%d",
            len(brands), matched, len(needs_review),
        )

        if needs_review:
            review_path = "needs_review.csv"
            with open(review_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = _csv.DictWriter(
                    f,
                    fieldnames=[
                        "brand_id", "name_cn",
                        "suggested_l1", "suggested_l2", "score",
                    ],
                )
                writer.writeheader()
                writer.writerows(needs_review)
            log.info(
                "需人工复核的匹配已输出: %s (%d 条)",
                review_path, len(needs_review),
            )

        if args.dry_run:
            log.info("DRY-RUN 模式: 未写库")
        else:
            conn.commit()
            log.info("已提交变更到数据库")
        conn.close()

        # Print deployment instructions
        print()
        print("=== 使用说明 ===")
        print("jufair-l2 爬取需在大陆 IP 环境执行。")
        print("请将 clean_brands.py + jufair_l2_crawler.py 复制到 Mac Mini")
        print("在 Mac Mini 运行:")
        print("  python scripts/clean_brands.py jufair-l2 --export jufair_cats.json")
        print("将输出的 jufair_cats.json 复制回本机")
        print("运行:")
        print("  python scripts/clean_brands.py jufair-l2 --import jufair_cats.json")
        return

    # Neither --export nor --import specified
    log.error("请指定 --export <输出路径> 或 --import <输入路径>")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="MWLAB Phase 05 品牌表清洗",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    for cmd_name, func, help_text in [
        ("name-en", cmd_name_en, "标准化英文名称"),
        ("industry", cmd_industry, "归并 industry_l1 到 6 个 MD 类别"),
        ("mds", cmd_mds, "从 Excel 标记 MD 自有品牌"),
        ("jufair-l2", cmd_jufair_l2, "爬取 jufair 分类并匹配 L2"),
    ]:
        p = sub.add_parser(cmd_name, help=help_text)
        p.add_argument("--db", default=str(DB_PATH), help="目标数据库路径")
        p.add_argument(
            "--dry-run", action="store_true",
            help="预览模式：不写库，仅打印将要执行的变更",
        )
        p.set_defaults(func=func)

    # jufair-l2 特有参数
    jufair_parser = sub.choices.get("jufair-l2")
    if jufair_parser:
        jufair_parser.add_argument(
            "--export", default="",
            help="输出 JSON 文件路径（在 Mac Mini 爬取后保存）",
        )
        jufair_parser.add_argument(
            "--import", dest="import_path", default="",
            help="输入 JSON 文件路径（从 Mac Mini 复制回本机后导入）",
        )
        jufair_parser.add_argument(
            "--threshold", type=float, default=0.80,
            help="模糊匹配阈值（默认 0.80）",
        )

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
