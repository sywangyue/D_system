#!/usr/bin/env python3
"""
tools/export_for_tagging.py — Phase 3b 导出待打标 Excel

用法:
  python tools/export_for_tagging.py --industry_l2 "机床" --status untagged
  python tools/export_for_tagging.py --industry_l2 "机床" --status all -o exports/batch.xlsx

依赖: openpyxl
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from datetime import date
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

DB_PATH = _REPO_ROOT / "mwlab.db"

TAGGABLE_FIELDS: dict[str, type] = {
    "competition_relation":  str,
    "mds_related":           str,
    "scale_score":           int,
    "is_international":      int,
    "is_ufi_certified":      int,
    "ma_potential":          int,
    "strategic_relevance":   int,
    "competitor_group":      str,
    "industry_l1":           str,
    "industry_l2":           str,
    "notes":                 str,
    "first_year":            int,
    "organizer":             str,
    "co_organizer":          str,
    "city":                  str,
    "frequency":             str,
    "website":               str,
}

# 固定列：主键 + 中文名/英文名（便于人工核对，不可打标）
_PREFIX_COLS = ["brand_id", "name_cn", "name_en"]
_TAG_COLUMNS = list(TAGGABLE_FIELDS.keys())
_ALL_EXPORT_COLS = _PREFIX_COLS + _TAG_COLUMNS


def _is_untagged_clause() -> str:
    """与 PRD「待打标」一致：竞争关系尚未填写。"""
    return "(TRIM(COALESCE(competition_relation, '')) = '')"


def _fetch_rows(
    conn: sqlite3.Connection,
    industry_l2: str,
    status: str,
) -> list[sqlite3.Row]:
    sql = (
        "SELECT " + ", ".join(_ALL_EXPORT_COLS) + " FROM exhibition_brand "
        "WHERE industry_l2 = ? "
    )
    params: list = [industry_l2]
    if status == "untagged":
        sql += f"AND {_is_untagged_clause()} "
    elif status != "all":
        raise ValueError(f"未知 --status: {status}")
    sql += "ORDER BY brand_id"
    return list(conn.execute(sql, params).fetchall())


def _apply_validations(ws) -> None:
    """对打标列（非前 3 列）附加下拉/列表校验。"""
    max_row = ws.max_row
    if max_row < 2:
        return
    # 列号 1-based：打标从第 4 列起
    first_tag_col = len(_PREFIX_COLS) + 1
    last_col = len(_ALL_EXPORT_COLS)

    def add_list(col_idx: int, options: list[str]) -> None:
        inner = ",".join(options)
        dv = DataValidation(
            type="list",
            formula1=f'"{inner}"',
            allow_blank=True,
        )
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        dv.add(f"{col_letter}2:{col_letter}{max_row}")
        ws.add_data_validation(dv)

    field_to_col = {ws.cell(1, c).value: c for c in range(1, last_col + 1)}

    cr_col = field_to_col.get("competition_relation")
    if cr_col:
        add_list(cr_col, ["是", "否"])

    for name, lo, hi in (
        ("scale_score", 1, 10),
        ("ma_potential", 1, 5),
        ("strategic_relevance", 1, 5),
    ):
        c = field_to_col.get(name)
        if c:
            add_list(c, [str(x) for x in range(lo, hi + 1)])

    for name in ("is_international", "is_ufi_certified"):
        c = field_to_col.get(name)
        if c:
            add_list(c, ["0", "1"])


def main() -> int:
    parser = argparse.ArgumentParser(description="导出待打标 Excel（Phase 3b）")
    parser.add_argument(
        "--industry_l2",
        required=True,
        help="筛选 exhibition_brand.industry_l2（必填）",
    )
    parser.add_argument(
        "--status",
        choices=("untagged", "all"),
        default="untagged",
        help="untagged=竞争关系为空；all=该品类全部",
    )
    parser.add_argument(
        "--output",
        "-o",
        default="",
        help="输出 .xlsx 路径（默认 exports/tagging_batch_YYYYMMDD.xlsx）",
    )
    parser.add_argument(
        "--db",
        default="",
        help="SQLite 路径（默认仓库根目录 mwlab.db）",
    )
    args = parser.parse_args()
    db_path = Path(args.db) if args.db else DB_PATH
    if not db_path.is_file():
        print(f"错误: 数据库不存在: {db_path}", file=sys.stderr)
        return 1

    out = Path(args.output) if args.output else (
        _REPO_ROOT / "exports" / f"tagging_batch_{date.today():%Y%m%d}.xlsx"
    )
    out.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(str(db_path), detect_types=sqlite3.PARSE_DECLTYPES)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        rows = _fetch_rows(conn, args.industry_l2.strip(), args.status)
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "tagging"

    for col_idx, key in enumerate(_ALL_EXPORT_COLS, start=1):
        cell = ws.cell(1, col_idx, key)
        if col_idx <= len(_PREFIX_COLS):
            cell.font = Font(bold=True)

    for r, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(_ALL_EXPORT_COLS, start=1):
            val = row[key]
            if val is None:
                ws.cell(r, col_idx, "")
            else:
                ws.cell(r, col_idx, val)

    _apply_validations(ws)
    ws.freeze_panes = "D2"
    wb.save(out)
    print(f"已导出 {len(rows)} 行 → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
