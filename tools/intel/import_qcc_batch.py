#!/usr/bin/env python3
"""
tools/intel/import_qcc_batch.py — 企查查批量导出 Excel → customer_prospect 表

用法:
  python3 tools/intel/import_qcc_batch.py /path/to/qcc_export.xlsx
  python3 tools/intel/import_qcc_batch.py /path/to/qcc_export.xlsx --dry-run

策略:
  - 直接映射 9 个字段到 DB 列
  - 剩余 30 个字段打包 JSON 存入 notes
  - 按 credit_code 去重: 存在则 UPDATE，不存在则 INSERT
  - source_type = 'qcc_batch'
"""
from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from pathlib import Path

import openpyxl

_SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = _SCRIPT_DIR.parent.parent
DB_PATH = PROJECT_ROOT / "data" / "mwlab.db"

# ── Excel 列索引 (0-based, 第2行为表头) ──────────────────────────────
_COL = {
    "search_key": 0,          # 原文件导入名称
    "company_name": 1,        # 系统匹配企业名称
    "company_status": 2,      # 登记状态
    "oper_name": 3,           # 法定代表人
    "reg_capital": 4,         # 注册资本
    "paid_capital": 5,        # 实缴资本
    "start_date": 6,          # 成立日期
    "credit_code": 7,         # 统一社会信用代码
    "address": 8,             # 企业地址
    "province": 9,            # 所属省份
    "city": 10,               # 所属城市
    "district": 11,           # 所属区县
    "phone": 12,              # 电话
    "more_phones": 13,        # 更多电话
    "email": 14,              # 邮箱
    "more_emails": 15,        # 更多邮箱
    "company_type": 16,       # 企业（机构）类型
    "tax_id": 17,             # 纳税人识别号
    "reg_no": 18,             # 注册号
    "org_code": 19,           # 组织机构代码
    "insured_count": 20,      # 参保人数
    "insured_year": 21,       # 参保人数所属年报
    "biz_term": 22,           # 营业期限
    "gb_industry_door": 23,   # 国标行业门类
    "gb_industry_big": 24,    # 国标行业大类
    "gb_industry_mid": 25,    # 国标行业中类
    "gb_industry_small": 26,  # 国标行业小类
    "qcc_industry_door": 27,  # 企查查行业门类
    "qcc_industry_big": 28,   # 企查查行业大类
    "qcc_industry_mid": 29,   # 企查查行业中类
    "qcc_industry_small": 30, # 企查查行业小类
    "company_scale": 31,      # 企业规模
    "former_name": 32,        # 曾用名
    "name_en": 33,            # 英文名
    "website": 34,            # 官网
    "mailing_addr": 35,       # 通信地址
    "description": 36,        # 企业简介
    "biz_scope": 37,          # 经营范围
    "reg_authority": 38,      # 登记机关
    "tax_qual": 39,           # 纳税人资质
    "latest_annual_year": 40, # 最新年报年份
}

# notes JSON 里存放的字段
_NOTES_KEYS = [
    "search_key", "reg_capital", "paid_capital",
    "province", "city", "district",
    "phone", "more_phones", "more_emails",
    "company_type", "tax_id", "org_code",
    "insured_count", "insured_year", "biz_term",
    "gb_industry_door", "gb_industry_big", "gb_industry_mid", "gb_industry_small",
    "qcc_industry_door", "qcc_industry_big", "qcc_industry_mid", "qcc_industry_small",
    "company_scale", "former_name", "name_en", "website", "mailing_addr",
    "description", "biz_scope", "reg_authority", "tax_qual", "latest_annual_year",
]


def _cell(row: tuple, key: str) -> str | None:
    """安全取单元格值，None → None, 空字符串 → None"""
    idx = _COL[key]
    val = row[idx] if idx < len(row) else None
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s != "-" else None


def _date(raw: str | None) -> str | None:
    """成立日期统一为 YYYY-MM-DD"""
    if not raw:
        return None
    # 企查查导出格式: "2005-03-30"
    raw = raw.strip()
    if len(raw) == 10 and raw[4] == "-":
        return raw
    # 尝试其他格式...
    return raw


def read_excel(path: str) -> list[dict]:
    """读取企查查批量导出 Excel，返回 dict 列表（第2行表头，第3行起数据）"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    for row_idx in range(3, ws.max_row + 1):  # 第1行声明, 第2行表头
        values = tuple(ws.cell(row=row_idx, column=c).value for c in range(1, 42))
        company_name = _cell(values, "company_name")
        if not company_name:
            continue  # 跳过空行

        record = {
            "company_name": company_name,
            "credit_code": _cell(values, "credit_code"),
            "oper_name": _cell(values, "oper_name"),
            "start_date": _date(_cell(values, "start_date")),
            "company_status": _cell(values, "company_status"),
            "reg_no": _cell(values, "reg_no"),
            "address": _cell(values, "address"),
            "email": _cell(values, "email"),
            "notes": _build_notes(values),
        }
        rows.append(record)

    wb.close()
    return rows


def _build_notes(values: tuple) -> str:
    """将额外字段序列化为 JSON"""
    extra = {}
    for key in _NOTES_KEYS:
        val = _cell(values, key)
        if val:
            extra[key] = val
    if not extra:
        return ""
    return json.dumps(extra, ensure_ascii=False)


def import_data(db_path: str, records: list[dict], dry_run: bool = False) -> dict:
    """执行导入，返回统计"""
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")

    inserted = 0
    updated = 0
    skipped = 0

    try:
        for r in records:
            credit_code = r["credit_code"]

            if credit_code:
                # 查重
                existing = conn.execute(
                    "SELECT id, source_type, qcc_key_no FROM customer_prospect WHERE credit_code = ? LIMIT 1",
                    (credit_code,),
                ).fetchone()
            else:
                # 无信用代码，按公司名查重
                existing = conn.execute(
                    "SELECT id, source_type, qcc_key_no FROM customer_prospect WHERE company_name = ? AND credit_code IS NULL LIMIT 1",
                    (r["company_name"],),
                ).fetchone()

            info = f'{r["company_name"]} ({credit_code or "无代码"})'

            if existing:
                if dry_run:
                    print(f"  [DRY-RUN] 跳过(已存在): {info}")
                    skipped += 1
                    continue

                # UPDATE — 保留 qcc_key_no 如果原来有（API 数据）
                eid, old_source, old_qcc = existing
                conn.execute(
                    """UPDATE customer_prospect SET
                        company_name = ?, credit_code = ?, oper_name = ?,
                        start_date = ?, company_status = ?, reg_no = ?,
                        address = ?, email = ?, notes = ?,
                        updated_at = datetime('now', 'localtime')
                    WHERE id = ?""",
                    (
                        r["company_name"], credit_code, r["oper_name"],
                        r["start_date"], r["company_status"], r["reg_no"],
                        r["address"], r["email"], r["notes"],
                        eid,
                    ),
                )
                updated += 1
            else:
                if dry_run:
                    print(f"  [DRY-RUN] INSERT: {info}")
                    inserted += 1
                    continue

                conn.execute(
                    """INSERT INTO customer_prospect
                        (source_type, company_name, credit_code, oper_name,
                         start_date, company_status, reg_no, address, email, notes)
                    VALUES ('qcc_search', ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        r["company_name"], credit_code, r["oper_name"],
                        r["start_date"], r["company_status"], r["reg_no"],
                        r["address"], r["email"], r["notes"],
                    ),
                )
                inserted += 1

        if not dry_run:
            conn.commit()

    finally:
        conn.close()

    return {"inserted": inserted, "updated": updated, "skipped": skipped, "total": len(records)}


def main():
    parser = argparse.ArgumentParser(description="企查查批量导出 Excel → customer_prospect")
    parser.add_argument("xlsx", help="企查查批量导出 .xlsx 路径")
    parser.add_argument("--db", default=str(DB_PATH), help="数据库路径")
    parser.add_argument("--dry-run", action="store_true", help="仅预览，不写入")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"错误: 文件不存在: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"📖 读取: {xlsx_path.name}")
    records = read_excel(str(xlsx_path))
    print(f"   共 {len(records)} 条有效记录\n")

    if args.dry_run:
        print("🔍 DRY-RUN 模式，不写入数据库:\n")

    stats = import_data(args.db, records, dry_run=args.dry_run)

    if args.dry_run:
        print(f"\n📊 DRY-RUN 统计: 新增 {stats['inserted']} | 跳过(已存在) {stats['skipped']} | 共 {stats['total']}")
    else:
        print(f"\n✅ 导入完成: 新增 {stats['inserted']} | 更新 {stats['updated']} | 共 {stats['total']}")

    # 验证
    if not args.dry_run:
        conn = sqlite3.connect(args.db)
        total = conn.execute("SELECT COUNT(*) FROM customer_prospect").fetchone()[0]
        batch = conn.execute(
            "SELECT COUNT(*) FROM customer_prospect WHERE notes != ''"
        ).fetchone()[0]
        conn.close()
        print(f"📊 DB 当前状态: customer_prospect 共 {total} 条 (其中含 notes 数据: {batch} 条)")


if __name__ == "__main__":
    main()
