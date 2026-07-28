#!/usr/bin/env python3
"""
export_exhibitions.py — 展会清单统一导出工具

合并自三个口径互不相同的旧脚本（AUDIT P2-20）：
    export_monthly.py       月度清单，city 取 b.city（品牌级），无同期展合并
    export_august.py        单月清单，含三级同期展合并，但月份写死
    export_2026_2027Q2.py   区间清单，city 取 e.city（届次级），无地区过滤

旧脚本对同一个月导出会给出三种不同结果。本脚本统一为：
  - city 一律取届次级（e.city），品牌级仅作兜底 —— 展会会巡回，届次城市才准确
  - 地区过滤和同期展合并都做成显式开关，不再隐含在脚本选择里

用法:
  python3 tools/export_exhibitions.py --month 2026-08
  python3 tools/export_exhibitions.py --from 2026-01-01 --to 2027-07-31 --region all
  python3 tools/export_exhibitions.py --month 2026-08 --no-merge -o /tmp/raw.xlsx
"""
from __future__ import annotations

import argparse
import os
import sqlite3
import sys
from calendar import monthrange
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from itertools import combinations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
from tools.geo_dict import CN_CITIES, CN_PROVINCES

DB_PATH = _REPO_ROOT / "data" / "mwlab.db"

# 非大陆地区：城市名 + 展馆关键词双重排除。
# 旧 export_monthly.py 只把港澳加进了城市集合，漏了展馆关键词，
# 于是 city 为空但展馆是「香港会议展览中心」的记录仍会混进「中国境内」清单（AUDIT P1-19）。
EXCLUDE_CITIES = {"台湾", "台北", "高雄", "台中", "台南", "桃园", "香港", "澳门", "港澳"}
EXCLUDE_VENUE_KEYWORDS = ("台湾", "台北", "香港", "澳门")

HEADERS = ["序号", "展会名称", "城市", "开展日期", "结束日期", "展馆",
           "面积(㎡)", "展商数", "观众数", "行业L1", "行业L2", "同期同馆子展"]
COL_WIDTHS = [6, 45, 10, 13, 13, 32, 12, 10, 12, 14, 14, 46]


# ── 日期范围 ────────────────────────────────────────────────────────────────
def resolve_range(args) -> tuple[str, str, str]:
    """返回 (date_from, date_to, label)。"""
    if args.month:
        y, m = int(args.month[:4]), int(args.month[5:7])
        last = monthrange(y, m)[1]
        return f"{y}-{m:02d}-01", f"{y}-{m:02d}-{last:02d}", f"{y}年{m}月"
    return args.date_from, args.date_to, f"{args.date_from}~{args.date_to}"


# ── 查询 ────────────────────────────────────────────────────────────────────
def fetch_rows(db_path: Path, date_from: str, date_to: str, mainland_only: bool) -> list[tuple]:
    conn = sqlite3.connect(str(db_path))
    try:
        rows = conn.execute("""
            SELECT b.name_cn,
                   COALESCE(NULLIF(e.city,''), b.city) AS city,
                   e.date_start, e.date_end, e.venue,
                   e.area_sqm, e.exhibitors_count, e.visitors_count,
                   b.industry_l1, b.industry_l2, b.brand_id
            FROM exhibition_edition e
            JOIN exhibition_brand b ON b.brand_id = e.brand_id
            WHERE e.date_start >= ? AND e.date_start <= ?
            ORDER BY e.date_start, e.venue, b.name_cn
        """, (date_from, date_to)).fetchall()
    finally:
        conn.close()

    if not mainland_only:
        return rows

    cn_set = set(CN_CITIES) | set(CN_PROVINCES)
    out = []
    for r in rows:
        city, venue = (r[1] or "").strip(), (r[4] or "").strip()
        if city in EXCLUDE_CITIES:
            continue
        if any(kw in venue for kw in EXCLUDE_VENUE_KEYWORDS):
            continue
        if city in cn_set:
            out.append(r)
    return out


# ── 同期同馆合并 ────────────────────────────────────────────────────────────
def common_prefix_len(strings: list[str]) -> int:
    if len(strings) < 2:
        return 0
    s0 = strings[0]
    for i, ch in enumerate(s0):
        for s in strings[1:]:
            if i >= len(s) or s[i] != ch:
                return i
    return len(s0)


def cluster_by_similarity(names: list[str], threshold: float = 0.45) -> list[list[int]]:
    """名称相似度连通分量聚类，返回索引分组。"""
    n = len(names)
    adj: dict[int, set[int]] = {i: set() for i in range(n)}
    for i, j in combinations(range(n), 2):
        if SequenceMatcher(None, names[i], names[j]).ratio() >= threshold:
            adj[i].add(j)
            adj[j].add(i)
    visited: set[int] = set()
    clusters: list[list[int]] = []
    for i in range(n):
        if i in visited:
            continue
        stack, comp = [i], []
        while stack:
            node = stack.pop()
            if node in visited:
                continue
            visited.add(node)
            comp.append(node)
            stack.extend(adj[node] - visited)
        clusters.append(comp)
    return clusters


def _fold(entries: list[tuple], ds: str, de: str, venue: str) -> tuple:
    """把同期同馆的多条折成一行，最短名为主名，其余进「同期同馆子展」列。"""
    ordered = sorted(entries, key=lambda x: len(x[0]))
    primary = ordered[0]
    aliases = [r[0] for r in ordered[1:]]
    areas = [r[5] for r in entries if r[5]]
    return (primary[0], primary[1], ds, de, venue,
            max(areas) if areas else None, None, None,
            primary[8], primary[9], primary[10], "；".join(aliases))


def merge_colocated(rows: list[tuple]) -> list[tuple]:
    """同 venue + 同起止日期的子展合并。

    前缀共享 ≥3 字 → 整组折叠；否则小群组(≤5)按名称相似度聚类；
    大群组（如上海新国际同期十余展）保持分开，避免误合并不相关展会。
    """
    groups: dict[tuple, list] = defaultdict(list)
    ungrouped: list[tuple] = []
    for r in rows:
        venue, ds, de = r[4], r[2], r[3]
        (groups[(ds, de, venue)] if (venue and ds and de) else ungrouped).append(r)

    merged: list[tuple] = []
    for (ds, de, venue), group in groups.items():
        if len(group) == 1:
            merged.append(group[0])
            continue
        names = [r[0] for r in group]
        if common_prefix_len(names) >= 3:
            merged.append(_fold(group, ds, de, venue))
        elif len(group) <= 5:
            for cl in cluster_by_similarity(names):
                picked = [group[i] for i in cl]
                merged.append(picked[0] if len(cl) == 1 else _fold(picked, ds, de, venue))
        else:
            merged.extend(group)

    merged.extend(ungrouped)
    merged.sort(key=lambda r: (r[2] or "", r[3] or "", r[1] or ""))
    return merged


# ── 写 Excel ────────────────────────────────────────────────────────────────
def write_xlsx(rows: list[tuple], out_path: Path, sheet_title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    hfont = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    dfont = Font(name="微软雅黑", size=10)
    bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(vertical="center", wrap_text=True)

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, center, bdr

    for i, r in enumerate(rows):
        aliases = r[11] if len(r) == 12 else ""
        name_cn, city, ds, de, venue, area, exh, vis, l1, l2 = r[0:10]
        vals = [i + 1, name_cn or "", city or "", ds or "", de or "", venue or "",
                area or "", exh or "", vis or "", l1 or "", l2 or "", aliases]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i + 2, column=col, value=v)
            c.font, c.border = dfont, bdr
            c.alignment = center if col in (1, 3, 4, 5, 7, 8, 9) else left

    for col, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ── CLI ─────────────────────────────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser(description="展会清单导出（统一口径）")
    period = ap.add_mutually_exclusive_group(required=True)
    period.add_argument("--month", metavar="YYYY-MM", help="按月导出")
    period.add_argument("--from", dest="date_from", metavar="YYYY-MM-DD", help="起始日期")
    ap.add_argument("--to", dest="date_to", metavar="YYYY-MM-DD", help="结束日期（配合 --from）")
    ap.add_argument("--region", choices=["mainland", "all"], default="mainland",
                    help="mainland=仅中国大陆（默认），all=不做地区过滤")
    ap.add_argument("--no-merge", action="store_true", help="不合并同期同馆子展")
    ap.add_argument("--db", default=str(DB_PATH), help="数据库路径（默认 data/mwlab.db）")
    ap.add_argument("-o", "--out", default="", help="输出 .xlsx 路径")
    args = ap.parse_args()

    if args.month and not __import__("re").fullmatch(r"\d{4}-\d{2}", args.month):
        ap.error(f"--month 格式须为 YYYY-MM，得到 {args.month!r}")
    if args.date_from and not args.date_to:
        ap.error("--from 必须配合 --to 使用")

    db_path = Path(args.db)
    if not db_path.is_file():
        print(f"错误: 数据库不存在 {db_path}", file=sys.stderr)
        return 1

    date_from, date_to, label = resolve_range(args)
    mainland = args.region == "mainland"

    rows = fetch_rows(db_path, date_from, date_to, mainland)
    if not rows:
        print(f"[WARN] {label} 无数据（region={args.region}）")
        return 1

    raw_count = len(rows)
    if not args.no_merge:
        rows = merge_colocated(rows)

    scope = "中国境内" if mainland else "全部地区"
    slug = args.month if args.month else f"{date_from}_{date_to}"
    out_path = Path(args.out) if args.out else _REPO_ROOT / "exports" / f"{slug}_{scope}展会清单.xlsx"

    write_xlsx(rows, out_path, f"{label}展会清单")

    print(f"[OK] {out_path}")
    print(f"  范围: {date_from} ~ {date_to} | 地区: {scope}")
    print(f"  合并前 {raw_count} 行 → 输出 {len(rows)} 行")
    return 0


if __name__ == "__main__":
    sys.exit(main())
