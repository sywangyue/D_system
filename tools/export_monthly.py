#!/usr/bin/env python3
"""
MWLAB 月度展会数据导出脚本
用法: python3 export_monthly.py [YYYY-MM]
不传参数时导出当月数据
"""
import sys
import sqlite3
from datetime import datetime, timedelta
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = "/Volumes/databoard/AI Project/D_dashboard/data/mwlab.db"
OUT_DIR = "/Volumes/databoard/AI Project/D_dashboard"

# 排除的非大陆城市/地区
EXCLUDE_CITIES = {'台北', '台湾'}
EXCLUDE_VENUE_KEYWORDS = ['台北', '台湾']


def get_month_range(ym: str | None = None):
    """解析月份，返回 (date_start, date_end, label)"""
    if ym:
        import re as _re  # [CORE-14] validate format before splitting
        if not _re.fullmatch(r'\d{4}-\d{2}', ym):
            print(f"用法: python3 export_monthly.py [YYYY-MM]\n错误: 月份格式无效 '{ym}'，须为 YYYY-MM（如 2026-06）", file=sys.stderr)
            sys.exit(1)
        year, month = map(int, ym.split('-'))
        if not (1 <= month <= 12):
            print(f"用法: python3 export_monthly.py [YYYY-MM]\n错误: 月份 {month} 超出范围 1-12", file=sys.stderr)
            sys.exit(1)
    else:
        now = datetime.now()
        year, month = now.year, now.month

    first_day = f"{year}-{month:02d}-01"
    last_day_num = monthrange(year, month)[1]
    last_day = f"{year}-{month:02d}-{last_day_num:02d}"
    label = f"{year}年{month}月"
    return first_day, last_day, label


def export(ym: str | None = None):
    date_from, date_to, month_label = get_month_range(ym)
    ym_slug = date_from[:7]  # YYYY-MM

    try:  # [CORE-14] managed connection — no leak on exception
        conn = sqlite3.connect(DB_PATH)
    except Exception as e:
        print(f"错误: 无法连接数据库 {DB_PATH}: {e}", file=sys.stderr)
        sys.exit(1)
    try:
        cur = conn.cursor()
        cur.execute("""
        SELECT b.name_cn, b.city, e.date_start, e.date_end, e.venue,
               e.area_sqm, e.exhibitors_count, e.visitors_count,
               b.industry_l1, b.industry_l2
        FROM exhibition_brand b
        JOIN exhibition_edition e ON b.brand_id = e.brand_id
        WHERE e.date_start >= ? AND e.date_start <= ?
          AND b.country_cn = '中国'
        ORDER BY e.date_start, b.city
    """, (date_from, date_to))
        rows = cur.fetchall()
    finally:
        conn.close()

    # 过滤台湾
    filtered = []
    for r in rows:
        city = (r[1] or '').strip()
        venue = (r[4] or '').strip()
        if city in EXCLUDE_CITIES:
            continue
        if any(kw in venue for kw in EXCLUDE_VENUE_KEYWORDS):
            continue
        filtered.append(r)

    if not filtered:
        print(f"[WARN] {month_label} 无中国境内展会数据")
        return None, month_label, 0

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_label}展会清单"

    headers = ["序号", "展会名称", "城市", "开展日期", "结束日期", "展馆",
               "面积(㎡)", "展商数", "观众数", "行业L1", "行业L2"]

    hfont = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    calign = Alignment(vertical="center", wrap_text=True)
    ccalign = Alignment(horizontal="center", vertical="center")
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"), bottom=Side(style="thin"))

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = halign
        c.border = bdr

    for i, r in enumerate(filtered):
        rn = i + 2
        name_cn, city, ds, de, venue, area, exh, vis, l1, l2 = r
        vals = [i+1, name_cn or '', city or '', ds or '', de or '',
                venue or '', area if area else '', exh if exh else '',
                vis if vis else '', l1 or '', l2 or '']
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=col, value=v)
            c.border = bdr
            c.alignment = ccalign if col in (1,3,4,5,7,8,9) else calign

    widths = [6, 48, 8, 14, 14, 34, 12, 10, 12, 16, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(filtered)+1}"

    filename = f"{ym_slug}_中国境内展会清单.xlsx"
    out_path = f"{OUT_DIR}/{filename}"
    wb.save(out_path)
    print(f"[OK] {out_path}  — {len(filtered)} 条记录")
    return out_path, month_label, len(filtered)


if __name__ == '__main__':
    ym = sys.argv[1] if len(sys.argv) > 1 else None
    path, label, count = export(ym)
    if path:
        print(f"COUNT:{count}")
        print(f"FILE:{path}")
