#!/usr/bin/env python3
"""
tools/import_tags.py — Phase 3b 从 Excel 写回打标

用法:
  python tools/import_tags.py --file exports/tagging_batch_20260506.xlsx --changed-by ops@example.com

空单元格视为「不修改该字段」。每条变更写入 manual_tag_history。
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path
from typing import Any

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from openpyxl import load_workbook

DB_PATH = _REPO_ROOT / "mwlab.db"

TAGGABLE_FIELDS: dict[str, type] = {
    "competition_relation":  str,
    "mds_related":           str,
    "scale_score":           int,
    "is_international":      int,
    "is_ufi_certified":      int,
    "ma_potential":          int,
    "strategic_relevance":   int,
    "competitor_group":      str,
    "industry_l1":           str,
    "industry_l2":           str,
    "notes":                 str,
    "first_year":            int,
    "organizer":             str,
    "co_organizer":          str,
    "city":                  str,
    "frequency":             str,
    "website":               str,
}

_ENUM_CONSTRAINTS: dict[str, set] = {
    "competition_relation": {"是", "否", ""},
}
_INT_RANGE_CONSTRAINTS: dict[str, tuple[int, int]] = {
    "scale_score":         (1, 10),
    "ma_potential":        (1, 5),
    "strategic_relevance": (1, 5),
    "is_international":    (0, 1),
    "is_ufi_certified":    (0, 1),
}


def validate_value(field_name: str, new_value: Any) -> tuple[bool, str]:
    expected_type = TAGGABLE_FIELDS.get(field_name)
    if expected_type is int:
        try:
            iv = int(new_value)
        except (ValueError, TypeError):
            return False, f"{field_name} 必须是整数，收到: {new_value!r}"
        if field_name in _INT_RANGE_CONSTRAINTS:
            lo, hi = _INT_RANGE_CONSTRAINTS[field_name]
            if not (lo <= iv <= hi):
                return False, f"{field_name} 必须在 [{lo}, {hi}] 范围内，收到: {iv}"
    if field_name in _ENUM_CONSTRAINTS:
        if str(new_value) not in _ENUM_CONSTRAINTS[field_name]:
            return False, (
                f"{field_name} 枚举值无效: {new_value!r}。"
                f"允许值: {_ENUM_CONSTRAINTS[field_name]}"
            )
    return True, ""


def _norm_cell(field: str, raw) -> str | int | None:
    """空白 → None（跳过），按字段类型强转。"""
    if raw is None:
        return None
    if isinstance(raw, str):
        s = raw.strip()
        if s == "":
            return None
        raw = s
    if TAGGABLE_FIELDS.get(field) is int:
        try:
            if isinstance(raw, float) and raw.is_integer():
                return int(raw)
            return int(raw)
        except (TypeError, ValueError):
            return raw
    return raw


def main() -> int:
    ap = argparse.ArgumentParser(description="从 Excel 导入打标（Phase 3b）")
    ap.add_argument("--file", "-f", required=True, help="export_for_tagging 产出的 .xlsx")
    ap.add_argument("--changed-by", required=True, help="写入 manual_tag_history.changed_by，须为邮箱")
    ap.add_argument("--reason", default="", help="可选：统一备注，写入每条 manual_tag_history.reason")
    ap.add_argument("--db", default="", help="SQLite 路径（默认 mwlab.db）")
    args = ap.parse_args()

    changed_by = args.changed_by.strip()
    if not changed_by or "@" not in changed_by:
        print("错误: --changed-by 须为有效邮箱", file=sys.stderr)
        return 1

    xlsx = Path(args.file)
    if not xlsx.is_file():
        print(f"错误: 文件不存在: {xlsx}", file=sys.stderr)
        return 1

    db_path = Path(args.db) if args.db else DB_PATH
    if not db_path.is_file():
        print(f"错误: 数据库不存在: {db_path}", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if not headers or not headers[0]:
        print("错误: 首行表头为空", file=sys.stderr)
        return 1

    col_index = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h}
    if "brand_id" not in col_index:
        print("错误: 表头缺少 brand_id", file=sys.stderr)
        return 1

    tag_cols = [f for f in TAGGABLE_FIELDS if f in col_index]

    conn = sqlite3.connect(str(db_path), detect_types=sqlite3.PARSE_DECLTYPES)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")

    ok = skip = err = 0
    errors: list[str] = []

    for r in range(2, ws.max_row + 1):
        bid_cell = ws.cell(r, col_index["brand_id"]).value
        if bid_cell is None or str(bid_cell).strip() == "":
            continue
        brand_id = str(bid_cell).strip()

        row_db = conn.execute(
            "SELECT * FROM exhibition_brand WHERE brand_id = ?", (brand_id,)
        ).fetchone()
        if not row_db:
            err += 1
            errors.append(f"行 {r}: 品牌不存在 {brand_id}")
            continue

        brand = dict(row_db)
        pending: list[tuple[str, Any, str, str]] = []
        row_err: str | None = None

        for field in tag_cols:
            col = col_index[field]
            raw = ws.cell(r, col).value
            coerced = _norm_cell(field, raw)
            if coerced is None:
                continue

            ok_v, msg = validate_value(field, coerced)
            if not ok_v:
                row_err = f"行 {r} {brand_id} 字段 {field}: {msg}"
                break

            if TAGGABLE_FIELDS[field] is int:
                new_v: str | int = int(coerced)
            else:
                new_v = str(coerced) if not isinstance(coerced, str) else coerced
            old_v = brand.get(field)
            old_s = "" if old_v is None else str(old_v)
            new_s = str(new_v)
            if old_s == new_s:
                continue
            pending.append((field, new_v, old_s, new_s))

        if row_err:
            err += 1
            errors.append(row_err)
            continue
        if not pending:
            skip += 1
            continue

        for field, new_v, old_s, new_s in pending:
            conn.execute(
                f"UPDATE exhibition_brand SET {field} = ?, "
                "updated_at = datetime('now','localtime') WHERE brand_id = ?",
                (new_v, brand_id),
            )
            conn.execute(
                """
                INSERT INTO manual_tag_history
                    (brand_id, field_name, old_value, new_value, changed_by, reason)
                VALUES (?,?,?,?,?,?)
                """,
                (brand_id, field, old_s, new_s, changed_by, args.reason or ""),
            )
        ok += 1

    conn.commit()
    conn.close()

    print(
        f"导入完成: 成功更新 {ok} 行（有字段变更） / "
        f"跳过 {skip} 行（无变更或仅空单元格） / 错误 {err} 条"
    )
    for line in errors[:50]:
        print(line, file=sys.stderr)
    if len(errors) > 50:
        print(f"... 另有 {len(errors) - 50} 条错误未显示", file=sys.stderr)
    return 1 if err else 0


if __name__ == "__main__":
    raise SystemExit(main())
