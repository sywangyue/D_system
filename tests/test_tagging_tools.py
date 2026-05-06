"""
tests/test_tagging_tools.py — Phase 3b 导出/导入工具测试

运行:
    python3 -m unittest tests/test_tagging_tools.py -v
"""
from __future__ import annotations

import os
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))

from openpyxl import Workbook

from schema.db import init_db


class TestTaggingTools(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self._tmp.close()
        self.db_path = Path(self._tmp.name)
        conn = init_db(self.db_path)
        conn.execute(
            "INSERT INTO exhibition_brand (brand_id, name_cn, name_en, industry_l2, "
            "competition_relation) VALUES (?,?,?,?,?)",
            ("EXPO-TST1", "测试展A", "Test A", "机床", ""),
        )
        conn.execute(
            "INSERT INTO exhibition_brand (brand_id, name_cn, name_en, industry_l2, "
            "competition_relation, strategic_relevance) VALUES (?,?,?,?,?,?)",
            ("EXPO-TST2", "测试展B", "Test B", "机床", "是", 3),
        )
        conn.commit()
        conn.close()

        self._xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        self._xlsx.close()
        self.xlsx_path = Path(self._xlsx.name)

    def tearDown(self):
        for p in (self.db_path, self.xlsx_path):
            try:
                os.unlink(p)
            except OSError:
                pass

    def test_export_untagged_only_tst1(self):
        from tools.export_for_tagging import _fetch_rows

        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        rows = _fetch_rows(conn, "机床", "untagged")
        conn.close()
        ids = [r["brand_id"] for r in rows]
        self.assertEqual(ids, ["EXPO-TST1"])

    def test_import_writes_history(self):
        from tag_api import TAGGABLE_FIELDS
        from tools import import_tags

        headers = ["brand_id", "name_cn", "name_en"] + list(TAGGABLE_FIELDS.keys())
        wb = Workbook()
        ws = wb.active
        for c, h in enumerate(headers, start=1):
            ws.cell(1, c, h)
        ws.cell(2, 1, "EXPO-TST1")
        ws.cell(2, headers.index("competition_relation") + 1, "是")
        wb.save(self.xlsx_path)

        prev_args = list(sys.argv)
        try:
            sys.argv = [
                "import_tags",
                "--file",
                str(self.xlsx_path),
                "--changed-by",
                "u@example.com",
                "--db",
                str(self.db_path),
            ]
            rc = import_tags.main()
        finally:
            sys.argv = prev_args
        self.assertEqual(rc, 0)

        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT competition_relation FROM exhibition_brand WHERE brand_id=?",
            ("EXPO-TST1",),
        ).fetchone()
        hist = conn.execute(
            "SELECT COUNT(*) FROM manual_tag_history WHERE brand_id=?",
            ("EXPO-TST1",),
        ).fetchone()[0]
        conn.close()
        self.assertEqual(dict(row)["competition_relation"], "是")
        self.assertGreaterEqual(hist, 1)


if __name__ == "__main__":
    unittest.main()
